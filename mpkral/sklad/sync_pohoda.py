#!/usr/bin/env python3
"""
M+P Král — Sync skladu z Pohody do Supabase
=============================================

Čte Excel export zásob z Pohody a nahrává do Supabase.

POSTUP:
1. V Pohodě: Sklady → Zásoby → Soubor → Datová komunikace → Export agendy → Excel 2007
2. Spusť: python sync_pohoda.py C:\cesta\k\Zásoby.xlsx
   nebo dvoj-klik na SYNC_SKLAD.bat (automaticky hledá Zásoby.xlsx na ploše)
"""

import json
import sys
import os
import urllib.request
from datetime import datetime

# ============================================
# CONFIG
# ============================================
SUPABASE_URL = "https://yqklnqmcjloxwdtifkjb.supabase.co"
SUPABASE_KEY = "sb_publishable_gZI-26x3NF4Dc5h-BuOpjg_3Y0n2Um-"

# Výchozí cesta k Excel exportu (Pohoda uloží na plochu)
DEFAULT_EXCEL = os.path.join(os.path.expanduser("~"), "Desktop", "Zásoby.xlsx")

# ============================================


def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}")


def read_stock_excel(path):
    """Čte zásoby z Excel exportu Pohody."""
    try:
        import openpyxl
    except ImportError:
        log("CHYBA: Nainstaluj openpyxl: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(path):
        log(f"CHYBA: Soubor nenalezen: {path}")
        log("Exportuj zásoby z Pohody: Soubor → Datová komunikace → Export agendy → Excel 2007")
        return []

    log(f"Čtu Excel: {path}")
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Najdi sloupce podle hlavičky
    header = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        val = (cell.value or '').strip()
        header[val] = cell.column - 1  # 0-indexed

    log(f"Sloupce: {list(header.keys())}")

    # Mapování sloupců Pohody
    col_kod = header.get('Kód', header.get('Kod', None))
    col_nazev = header.get('Název', header.get('Nazev', None))
    col_barcode = header.get('Čárkód', header.get('Carkod', header.get('EAN', None)))
    col_qty = header.get('Stav zásoby', header.get('Stav zasoby', header.get('Množství', None)))

    if col_kod is None or col_nazev is None:
        log(f"CHYBA: Nenalezeny sloupce Kód/Název v hlavičce: {list(header.keys())}")
        return []

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        kod = (str(row[col_kod]) if row[col_kod] else '').strip()
        if not kod:
            continue

        nazev = (str(row[col_nazev]) if row[col_nazev] else '').strip()
        carovy_kod = (str(row[col_barcode]) if col_barcode is not None and row[col_barcode] else None)
        mnozstvi = int(float(row[col_qty])) if col_qty is not None and row[col_qty] else 0

        items.append({
            'kod': kod,
            'nazev': nazev,
            'carovy_kod': carovy_kod,
            'mnozstvi': mnozstvi,
        })

    log(f"Načteno {len(items)} položek")
    return items


def upload_to_supabase(items):
    """Nahraje položky do Supabase (smaže staré, vloží nové)."""
    log(f"Nahrávám {len(items)} položek do Supabase...")

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }

    # 1. Smaž všechna stará data
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/sklad?id=gt.0",
        headers={**headers, "Prefer": "return=minimal"},
        method='DELETE'
    )
    urllib.request.urlopen(req)
    log("  Stará data smazána")

    # 2. Vlož nová data
    payload = json.dumps(items).encode()
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/sklad",
        data=payload,
        headers={**headers, "Prefer": "return=minimal"},
        method='POST'
    )
    urllib.request.urlopen(req)
    log(f"  {len(items)} položek nahráno")
    log("Upload dokončen!")
    return True


def main():
    print("=" * 50)
    log("M+P Král — Sync skladu Pohoda → Web")
    print("=" * 50)

    # Cesta k Excel souboru
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        excel_path = DEFAULT_EXCEL

    # 1. Načti data z Excelu
    items = read_stock_excel(excel_path)
    if not items:
        log("Žádné položky k nahrání!")
        input("Stiskni Enter pro zavření...")
        return

    # 2. Zobraz co se nahraje
    print()
    log("Položky k nahrání:")
    for i in items:
        print(f"  {i['kod']:30s} | {i['nazev']:30s} | {i['mnozstvi']}")
    print()

    # 3. Nahraj do Supabase
    try:
        upload_to_supabase(items)
    except Exception as e:
        log(f"CHYBA při nahrávání: {e}")
        input("Stiskni Enter pro zavření...")
        return

    print("=" * 50)
    log("Hotovo! Data jsou na webu.")
    input("Stiskni Enter pro zavření...")


if __name__ == "__main__":
    main()
